from __future__ import annotations

import csv
import json
import os
import shutil
import uuid
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

from bson import ObjectId
from pymongo import UpdateOne

from api.extensions import mongo


DEFAULT_YEAR = 2025
DEFAULT_CURRENCY = "VES"


def _now_utc() -> datetime:
    return datetime.now(timezone.utc)


def _to_object_id(value: str) -> Optional[ObjectId]:
    try:
        return ObjectId(str(value).strip())
    except Exception:
        return None


def _is_truthy(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    return str(value).strip().lower() in {"1", "true", "t", "yes", "si"}


def _clean_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


@dataclass
class SeedStats:
    accounts: int = 0
    units: int = 0
    funding_sources: int = 0
    budget_categories: int = 0

    def as_dict(self) -> Dict[str, int]:
        return {
            "accounts": self.accounts,
            "units": self.units,
            "funding_sources": self.funding_sources,
            "budget_categories": self.budget_categories,
        }


class AccountingIndexes:
    _created = False

    @classmethod
    def ensure_indexes(cls) -> None:
        if cls._created:
            return
        db = mongo.db
        db.master_accounts.create_index([("year", 1), ("code", 1)], unique=True)
        db.master_accounts.create_index([("year", 1), ("group", 1)])
        db.master_accounts.create_index([("year", 1), ("parent_code", 1)])

        db.master_units.create_index([("year", 1), ("code", 1)], unique=True)

        db.master_funding_sources.create_index([("year", 1), ("code", 1)], unique=True)
        db.master_budget_categories.create_index([("year", 1), ("code", 1)], unique=True)

        db.account_scope_state.create_index(
            [("year", 1), ("scopeType", 1), ("scopeId", 1), ("accountCode", 1)], unique=True
        )
        db.account_scope_state.create_index([("year", 1), ("scopeType", 1), ("scopeId", 1)])
        db.account_scope_state.create_index([("year", 1), ("accountCode", 1)])

        db.ledger_movements.create_index([("year", 1), ("scopeType", 1), ("scopeId", 1), ("createdAt", -1)])
        db.ledger_movements.create_index([("year", 1), ("accountCode", 1)])
        db.ledger_movements.create_index([("year", 1), ("reference.kind", 1), ("reference.id", 1)])

        db.departamentos.create_index([("accountingUnitCode", 1)], sparse=True)
        cls._created = True


class AccountCatalogService:
    @staticmethod
    def search(year: int, q: str = "", group: Optional[str] = None, limit: int = 50) -> List[Dict[str, Any]]:
        AccountingIndexes.ensure_indexes()
        query: Dict[str, Any] = {"year": int(year)}
        if group:
            query["group"] = group.upper()
        if q:
            safe_q = str(q).strip()
            query["$or"] = [{"code": {"$regex": safe_q}}, {"description": {"$regex": safe_q, "$options": "i"}}]

        cursor = mongo.db.master_accounts.find(query, {"_id": 0}).sort("code", 1).limit(max(1, min(limit, 500)))
        return list(cursor)

    @staticmethod
    def tree(year: int, group: Optional[str] = None) -> List[Dict[str, Any]]:
        AccountingIndexes.ensure_indexes()
        query: Dict[str, Any] = {"year": int(year)}
        if group:
            query["group"] = group.upper()

        accounts = list(mongo.db.master_accounts.find(query, {"_id": 0}).sort("code", 1))
        return _build_tree(accounts)

    @staticmethod
    def consolidated_totals(year: int, scope_type: Optional[str] = None, scope_id: Optional[str] = None) -> Dict[str, Any]:
        AccountingIndexes.ensure_indexes()
        match: Dict[str, Any] = {"year": int(year)}
        if scope_type:
            match["scopeType"] = scope_type
        if scope_id:
            match["scopeId"] = scope_id

        pipeline = [
            {"$match": match},
            {
                "$group": {
                    "_id": "$accountCode",
                    "totalBalance": {"$sum": "$balance"},
                    "totalMovements": {"$sum": "$movementsCount"},
                }
            },
            {"$sort": {"_id": 1}},
        ]

        totals_rows = list(mongo.db.account_scope_state.aggregate(pipeline))
        rows = [
            {
                "accountCode": row["_id"],
                "balance": row.get("totalBalance", 0),
                "movementsCount": row.get("totalMovements", 0),
            }
            for row in totals_rows
        ]

        accounts = list(mongo.db.master_accounts.find({"year": int(year)}, {"_id": 0, "code": 1, "parent_code": 1, "description": 1}))
        by_code = {acc["code"]: acc for acc in accounts}

        def find_root(code: str) -> str:
            current = by_code.get(code)
            seen = set()
            while current and current.get("parent_code") and current["parent_code"] in by_code and current["code"] not in seen:
                seen.add(current["code"])
                current = by_code[current["parent_code"]]
            return current["code"] if current else code

        root_agg: Dict[str, float] = {}
        for row in rows:
            root_code = find_root(row["accountCode"])
            root_agg[root_code] = root_agg.get(root_code, 0) + float(row.get("balance", 0))

        totals_by_root = [
            {
                "rootCode": code,
                "description": by_code.get(code, {}).get("description", "N/A"),
                "balance": balance,
            }
            for code, balance in sorted(root_agg.items())
        ]

        return {"year": int(year), "totalsByAccount": rows, "totalsByRoot": totals_by_root}


class AccountScopeService:
    @staticmethod
    def get_scope_accounts(
        year: int,
        scope_type: str,
        scope_id: str,
        group: Optional[str] = None,
        assigned_only: bool = False,
        include_zero: bool = True,
    ) -> Dict[str, Any]:
        AccountingIndexes.ensure_indexes()
        accounts_query: Dict[str, Any] = {"year": int(year)}
        if group:
            accounts_query["group"] = group.upper()

        accounts = list(mongo.db.master_accounts.find(accounts_query, {"_id": 0}).sort("code", 1))
        states_cursor = mongo.db.account_scope_state.find(
            {"year": int(year), "scopeType": scope_type, "scopeId": scope_id},
            {"_id": 0, "accountCode": 1, "balance": 1, "movementsCount": 1, "lastMovementAt": 1},
        )
        state_by_code = {row["accountCode"]: row for row in states_cursor}
        all_assigned_codes = set(state_by_code.keys())

        by_code: Dict[str, Dict[str, Any]] = {}

        merged = []
        for account in accounts:
            state = state_by_code.get(account["code"], {})
            has_state = account["code"] in all_assigned_codes
            item = {
                **account,
                "balance": state.get("balance", 0),
                "movementsCount": state.get("movementsCount", 0),
                "lastMovementAt": state.get("lastMovementAt"),
                "hasState": has_state,
            }
            by_code[item["code"]] = item
            merged.append(
                item
            )

        if assigned_only:
            visible_codes = set(all_assigned_codes)
        else:
            visible_codes = set(by_code.keys())

        if not include_zero:
            non_zero_codes = {
                code
                for code in visible_codes
                if float(by_code.get(code, {}).get("balance", 0) or 0) != 0
                or float(by_code.get(code, {}).get("movementsCount", 0) or 0) > 0
            }

            # Preserve parent chain so the tree remains navigable.
            with_ancestors = set(non_zero_codes)
            for code in list(non_zero_codes):
                current = by_code.get(code)
                while current and current.get("parent_code"):
                    parent_code = current["parent_code"]
                    parent = by_code.get(parent_code)
                    if not parent or parent_code in with_ancestors:
                        break
                    with_ancestors.add(parent_code)
                    current = parent
            visible_codes = with_ancestors

        filtered_items = [item for item in merged if item["code"] in visible_codes]
        total_visible = len(filtered_items)
        total_balance_visible = sum(float(item.get("balance", 0) or 0) for item in filtered_items)

        return {
            "year": int(year),
            "scopeType": scope_type,
            "scopeId": scope_id,
            "tree": _build_tree(filtered_items),
            "meta": {
                "assignedOnly": bool(assigned_only),
                "includeZero": bool(include_zero),
                "totalAssigned": len(all_assigned_codes),
                "totalVisible": total_visible,
                "totalBalanceVisible": total_balance_visible,
            },
        }

    @staticmethod
    def init_scope(year: int, scope_type: str, scope_id: str, mode: str = "detail_only") -> Dict[str, Any]:
        AccountingIndexes.ensure_indexes()
        mode_value = (mode or "detail_only").strip()

        query: Dict[str, Any] = {"year": int(year)}
        if mode_value == "detail_only":
            query["is_header"] = False
        elif mode_value.startswith("group:"):
            group = mode_value.split(":", 1)[1].strip().upper()
            if group:
                query["group"] = group
        elif mode_value == "all":
            pass
        else:
            raise ValueError("mode inválido. Use detail_only, all o group:EGRESO")

        accounts = list(mongo.db.master_accounts.find(query, {"_id": 0, "code": 1}))
        if not accounts:
            return {"inserted": 0, "mode": mode_value, "scopeType": scope_type, "scopeId": scope_id, "year": int(year)}

        now = _now_utc()
        ops = []
        for account in accounts:
            ops.append(
                UpdateOne(
                    {
                        "year": int(year),
                        "scopeType": scope_type,
                        "scopeId": scope_id,
                        "accountCode": account["code"],
                    },
                    {
                        "$setOnInsert": {
                            "balance": 0,
                            "movementsCount": 0,
                            "lastMovementAt": None,
                            "createdAt": now,
                        },
                        "$set": {"updatedAt": now},
                    },
                    upsert=True,
                )
            )

        result = mongo.db.account_scope_state.bulk_write(ops, ordered=False)
        inserted = (result.upserted_count or 0)
        return {"inserted": int(inserted), "mode": mode_value, "scopeType": scope_type, "scopeId": scope_id, "year": int(year)}

    @staticmethod
    def create_movement(
        *,
        year: int,
        scope_type: str,
        scope_id: str,
        account_code: str,
        movement_type: str,
        amount: float,
        description: str,
        reference: Optional[Dict[str, Any]],
        created_by: str,
        allow_negative: bool,
    ) -> Dict[str, Any]:
        AccountingIndexes.ensure_indexes()
        if amount <= 0:
            raise ValueError("amount debe ser mayor que 0")

        movement_type = (movement_type or "").strip().lower()
        if movement_type not in {"debit", "credit"}:
            raise ValueError("type debe ser debit o credit")

        account = mongo.db.master_accounts.find_one({"year": int(year), "code": account_code})
        if not account:
            raise ValueError("La cuenta contable no existe para el año indicado")

        # En alcance global se permite cargar saldo en cuentas titular para consolidado base.
        if account.get("is_header") and scope_type != "global":
            raise ValueError("No se permiten movimientos sobre cuentas titular. Use cuentas detalle")

        delta = float(amount) if movement_type == "debit" else -float(amount)
        movement_doc = {
            "year": int(year),
            "scopeType": scope_type,
            "scopeId": scope_id,
            "accountCode": account_code,
            "type": movement_type,
            "amount": float(amount),
            "currency": DEFAULT_CURRENCY,
            "description": description or "",
            "reference": reference or {},
            "createdBy": created_by,
            "createdAt": _now_utc(),
        }

        state_filter = {
            "year": int(year),
            "scopeType": scope_type,
            "scopeId": scope_id,
            "accountCode": account_code,
        }

        existing = mongo.db.account_scope_state.find_one(state_filter, {"balance": 1})
        current_balance = float(existing.get("balance", 0)) if existing else 0.0
        new_balance = current_balance + delta
        if not allow_negative and new_balance < 0:
            raise ValueError("El movimiento deja saldo negativo y la política actual lo prohíbe")

        now = _now_utc()

        # Intento transaccional (si el clúster lo soporta). Si falla, fallback lógico.
        client = mongo.cx
        try:
            with client.start_session() as session:
                def _txn(sess):
                    mongo.db.ledger_movements.insert_one(movement_doc, session=sess)
                    mongo.db.account_scope_state.update_one(
                        state_filter,
                        {
                            "$setOnInsert": {"createdAt": now},
                            "$inc": {"balance": delta, "movementsCount": 1},
                            "$set": {"lastMovementAt": now, "updatedAt": now},
                        },
                        upsert=True,
                        session=sess,
                    )

                session.with_transaction(_txn)
        except Exception:
            mongo.db.ledger_movements.insert_one(movement_doc)
            mongo.db.account_scope_state.update_one(
                state_filter,
                {
                    "$setOnInsert": {"createdAt": now},
                    "$inc": {"balance": delta, "movementsCount": 1},
                    "$set": {"lastMovementAt": now, "updatedAt": now},
                },
                upsert=True,
            )

        state = mongo.db.account_scope_state.find_one(state_filter, {"_id": 0}) or {}
        movement_doc.pop("_id", None)
        return {"movement": movement_doc, "state": state}

    @staticmethod
    def transfer_between_accounts(
        *,
        year: int,
        scope_type: Optional[str] = None,
        scope_id: Optional[str] = None,
        from_scope_type: Optional[str] = None,
        from_scope_id: Optional[str] = None,
        to_scope_type: Optional[str] = None,
        to_scope_id: Optional[str] = None,
        from_account_code: str,
        to_account_code: str,
        amount: float,
        description: str,
        reference: Optional[Dict[str, Any]],
        created_by: str,
        allow_negative: bool,
    ) -> Dict[str, Any]:
        AccountingIndexes.ensure_indexes()
        valid_scopes = {"department", "project", "global"}
        if amount <= 0:
            raise ValueError("amount debe ser mayor que 0")

        resolved_from_scope_type = (from_scope_type or scope_type or "").strip()
        resolved_to_scope_type = (to_scope_type or scope_type or resolved_from_scope_type or "").strip()
        resolved_from_scope_id = str(from_scope_id or scope_id or "").strip()
        resolved_to_scope_id = str(to_scope_id or scope_id or "").strip()

        if resolved_from_scope_type not in valid_scopes or resolved_to_scope_type not in valid_scopes:
            raise ValueError("scopeType debe ser department, project o global")
        if resolved_from_scope_type == "global" and not resolved_from_scope_id:
            resolved_from_scope_id = "global"
        if resolved_to_scope_type == "global" and not resolved_to_scope_id:
            resolved_to_scope_id = "global"
        if not resolved_from_scope_id or not resolved_to_scope_id:
            raise ValueError("scopeId es requerido para ambos extremos de la transferencia")
        if (
            from_account_code == to_account_code
            and resolved_from_scope_type == resolved_to_scope_type
            and resolved_from_scope_id == resolved_to_scope_id
        ):
            raise ValueError("La cuenta origen y destino deben ser distintas")

        source = mongo.db.master_accounts.find_one({"year": int(year), "code": from_account_code})
        target = mongo.db.master_accounts.find_one({"year": int(year), "code": to_account_code})
        if not source or not target:
            raise ValueError("La cuenta origen o destino no existe para el año indicado")
        if source.get("is_header") or target.get("is_header"):
            raise ValueError("Las transferencias requieren cuentas detalle, no cuentas titular")

        source_filter = {
            "year": int(year),
            "scopeType": resolved_from_scope_type,
            "scopeId": resolved_from_scope_id,
            "accountCode": from_account_code,
        }
        target_filter = {
            "year": int(year),
            "scopeType": resolved_to_scope_type,
            "scopeId": resolved_to_scope_id,
            "accountCode": to_account_code,
        }

        source_state = mongo.db.account_scope_state.find_one(source_filter, {"balance": 1}) or {"balance": 0}
        source_balance = float(source_state.get("balance", 0))
        if not allow_negative and (source_balance - float(amount)) < 0:
            raise ValueError("Saldo insuficiente en la cuenta origen")

        transfer_id = str(uuid.uuid4())
        now = _now_utc()
        transfer_ref = {
            "kind": "transfer",
            "id": transfer_id,
            "fromScopeType": resolved_from_scope_type,
            "fromScopeId": resolved_from_scope_id,
            "toScopeType": resolved_to_scope_type,
            "toScopeId": resolved_to_scope_id,
            **(reference or {}),
        }

        source_movement = {
            "year": int(year),
            "scopeType": resolved_from_scope_type,
            "scopeId": resolved_from_scope_id,
            "accountCode": from_account_code,
            "type": "credit",
            "amount": float(amount),
            "currency": DEFAULT_CURRENCY,
            "description": description or f"Transferencia a {to_account_code}",
            "reference": transfer_ref,
            "createdBy": created_by,
            "createdAt": now,
        }

        target_movement = {
            "year": int(year),
            "scopeType": resolved_to_scope_type,
            "scopeId": resolved_to_scope_id,
            "accountCode": to_account_code,
            "type": "debit",
            "amount": float(amount),
            "currency": DEFAULT_CURRENCY,
            "description": description or f"Transferencia desde {from_account_code}",
            "reference": transfer_ref,
            "createdBy": created_by,
            "createdAt": now,
        }

        client = mongo.cx
        try:
            with client.start_session() as session:
                def _txn(sess):
                    mongo.db.ledger_movements.insert_many([source_movement, target_movement], session=sess)
                    mongo.db.account_scope_state.update_one(
                        source_filter,
                        {
                            "$setOnInsert": {"createdAt": now},
                            "$inc": {"balance": -float(amount), "movementsCount": 1},
                            "$set": {"lastMovementAt": now, "updatedAt": now},
                        },
                        upsert=True,
                        session=sess,
                    )
                    mongo.db.account_scope_state.update_one(
                        target_filter,
                        {
                            "$setOnInsert": {"createdAt": now},
                            "$inc": {"balance": float(amount), "movementsCount": 1},
                            "$set": {"lastMovementAt": now, "updatedAt": now},
                        },
                        upsert=True,
                        session=sess,
                    )

                session.with_transaction(_txn)
        except Exception:
            mongo.db.ledger_movements.insert_many([source_movement, target_movement])
            mongo.db.account_scope_state.update_one(
                source_filter,
                {
                    "$setOnInsert": {"createdAt": now},
                    "$inc": {"balance": -float(amount), "movementsCount": 1},
                    "$set": {"lastMovementAt": now, "updatedAt": now},
                },
                upsert=True,
            )
            mongo.db.account_scope_state.update_one(
                target_filter,
                {
                    "$setOnInsert": {"createdAt": now},
                    "$inc": {"balance": float(amount), "movementsCount": 1},
                    "$set": {"lastMovementAt": now, "updatedAt": now},
                },
                upsert=True,
            )

        source_new = mongo.db.account_scope_state.find_one(source_filter, {"_id": 0}) or {}
        target_new = mongo.db.account_scope_state.find_one(target_filter, {"_id": 0}) or {}
        return {
            "transferId": transfer_id,
            "fromScopeType": resolved_from_scope_type,
            "fromScopeId": resolved_from_scope_id,
            "toScopeType": resolved_to_scope_type,
            "toScopeId": resolved_to_scope_id,
            "fromAccountCode": from_account_code,
            "toAccountCode": to_account_code,
            "amount": float(amount),
            "sourceState": source_new,
            "targetState": target_new,
        }


class SeedService:
    def __init__(self, base_dir: Optional[Path] = None):
        root = Path(base_dir or Path(__file__).resolve().parents[2])
        self.repo_root = root
        self.data_dir = root / "data" / "contabilidad" / "2025"
        self.downloads_dir = Path.home() / "Downloads"

    def seed(self, year: int = DEFAULT_YEAR, force: bool = False, dry_run: bool = False) -> Dict[str, Any]:
        self._ensure_local_data_files()

        accounts = self._load_accounts()
        units = self._load_units()
        sources = self._load_funding_sources()
        categories = self._load_budget_categories()

        if dry_run:
            return {
                "year": int(year),
                "dryRun": True,
                "counts": {
                    "accounts": len(accounts),
                    "units": len(units),
                    "funding_sources": len(sources),
                    "budget_categories": len(categories),
                },
            }

        AccountingIndexes.ensure_indexes()
        db = mongo.db
        if force:
            db.master_accounts.delete_many({"year": int(year)})
            db.master_units.delete_many({"year": int(year)})
            db.master_funding_sources.delete_many({"year": int(year)})
            db.master_budget_categories.delete_many({"year": int(year)})

        stats = SeedStats()
        stats.accounts = self._bulk_upsert(
            db.master_accounts,
            [
                {
                    "year": int(year),
                    "code": row["code"],
                    "description": row["description"],
                    "group": row["group"],
                    "is_header": row["is_header"],
                    "level": row["level"],
                    "parent_code": row["parent_code"],
                    "updatedAt": _now_utc(),
                }
                for row in accounts
            ],
            key_fields=("year", "code"),
        )

        stats.units = self._bulk_upsert(
            db.master_units,
            [
                {
                    "year": int(year),
                    "code": row["code"],
                    "description": row["description"],
                    "level": row["level"],
                    "parent_code": row["parent_code"],
                }
                for row in units
            ],
            key_fields=("year", "code"),
        )

        stats.funding_sources = self._bulk_upsert(
            db.master_funding_sources,
            [{"year": int(year), "code": row["code"], "description": row["description"]} for row in sources],
            key_fields=("year", "code"),
        )

        stats.budget_categories = self._bulk_upsert(
            db.master_budget_categories,
            [{"year": int(year), "code": row["code"], "description": row["description"]} for row in categories],
            key_fields=("year", "code"),
        )

        return {"year": int(year), "dryRun": False, "force": bool(force), "counts": stats.as_dict()}

    def sync_departments_from_units(self, year: int = DEFAULT_YEAR) -> Dict[str, Any]:
        units = list(mongo.db.master_units.find({"year": int(year)}, {"_id": 0}).sort("code", 1))
        if not units:
            return {"created": 0, "updated": 0, "mapped": 0}

        created = 0
        updated = 0
        unit_to_dept_id: Dict[str, ObjectId] = {}

        for unit in units:
            now = _now_utc()
            filter_query = {"$or": [{"accountingUnitCode": unit["code"]}, {"codigo": unit["code"]}]}
            existing = mongo.db.departamentos.find_one(filter_query)
            payload = {
                "nombre": unit["description"],
                "descripcion": f"Unidad ejecutora {unit['code']} (contabilidad {year})",
                "codigo": unit["code"],
                "accountingUnitCode": unit["code"],
                "updated_at": now,
            }

            if existing:
                mongo.db.departamentos.update_one({"_id": existing["_id"]}, {"$set": payload})
                unit_to_dept_id[unit["code"]] = existing["_id"]
                updated += 1
            else:
                to_insert = {**payload, "fecha_creacion": now, "activo": True}
                insert_result = mongo.db.departamentos.insert_one(to_insert)
                unit_to_dept_id[unit["code"]] = insert_result.inserted_id
                created += 1

        mapped = 0
        for unit in units:
            parent_code = unit.get("parent_code")
            if not parent_code:
                continue
            dept_id = unit_to_dept_id.get(unit["code"])
            parent_id = unit_to_dept_id.get(parent_code)
            if dept_id and parent_id:
                mongo.db.departamentos.update_one({"_id": dept_id}, {"$set": {"parentDepartmentId": parent_id}})
                mapped += 1

        return {"created": created, "updated": updated, "mapped": mapped}

    def _bulk_upsert(self, collection, rows: List[Dict[str, Any]], key_fields: Tuple[str, ...]) -> int:
        if not rows:
            return 0

        now = _now_utc()
        ops = []
        for row in rows:
            key = {field: row[field] for field in key_fields}
            set_payload = {k: v for k, v in row.items() if k not in key_fields}
            set_payload["updatedAt"] = now
            ops.append(
                UpdateOne(
                    key,
                    {
                        "$set": set_payload,
                        "$setOnInsert": {"createdAt": now},
                    },
                    upsert=True,
                )
            )

        result = collection.bulk_write(ops, ordered=False)
        return int((result.upserted_count or 0) + (result.modified_count or 0))

    def _ensure_local_data_files(self) -> None:
        self.data_dir.mkdir(parents=True, exist_ok=True)

        required_files = [
            "contabilidad_2025_accounts.csv",
            "contabilidad_2025_unidades_ejecutoras.csv",
            "contabilidad_2025_fuentes_financiamiento.csv",
            "contabilidad_2025_categoria_presupuestaria.csv",
        ]

        missing = [name for name in required_files if not (self.data_dir / name).exists()]
        if not missing:
            self._ensure_accounts_json()
            return

        # 1) Fallback directo a CSV en Downloads (si existen)
        copied = 0
        for name in list(missing):
            src = self.downloads_dir / name
            dst = self.data_dir / name
            if src.exists():
                shutil.copy2(src, dst)
                copied += 1

        missing = [name for name in required_files if not (self.data_dir / name).exists()]
        if not missing:
            self._ensure_accounts_json()
            return

        # 2) Si faltan archivos, intentar regenerar desde XLSX
        xlsx_candidates = [
            self.data_dir / "TABLAS DE CONTABILIDAD AÑO 2025 (Sistema).xlsx",
            self.downloads_dir / "TABLAS DE CONTABILIDAD AÑO 2025 (Sistema).xlsx",
        ]
        xlsx_path = next((p for p in xlsx_candidates if p.exists()), None)
        if not xlsx_path:
            raise FileNotFoundError(
                "No se encontraron CSV ni el Excel base para regenerar datos de contabilidad 2025"
            )

        self._regenerate_csv_from_xlsx(xlsx_path)
        missing = [name for name in required_files if not (self.data_dir / name).exists()]
        if missing:
            raise RuntimeError(f"No se pudieron regenerar todos los archivos requeridos: {missing}")

        self._ensure_accounts_json()

    def _regenerate_csv_from_xlsx(self, xlsx_path: Path) -> None:
        try:
            from openpyxl import load_workbook
        except Exception as exc:
            raise RuntimeError(
                "Faltó dependencia openpyxl para regenerar CSV desde XLSX"
            ) from exc

        wb = load_workbook(str(xlsx_path), data_only=True)
        sheet_map = {
            "contabilidad_2025_accounts.csv": ["accounts", "cuentas", "master_accounts"],
            "contabilidad_2025_unidades_ejecutoras.csv": ["unidades", "unidades ejecutoras", "units"],
            "contabilidad_2025_fuentes_financiamiento.csv": ["fuentes", "fuentes financiamiento", "funding"],
            "contabilidad_2025_categoria_presupuestaria.csv": ["categoria", "categorias", "categoria presupuestaria"],
        }

        lower_names = {ws.title.lower(): ws for ws in wb.worksheets}

        for out_name, aliases in sheet_map.items():
            worksheet = None
            for alias in aliases:
                for sheet_name, ws in lower_names.items():
                    if alias in sheet_name:
                        worksheet = ws
                        break
                if worksheet:
                    break

            if worksheet is None:
                continue

            rows = list(worksheet.iter_rows(values_only=True))
            if not rows:
                continue
            headers = [str(h).strip() if h is not None else "" for h in rows[0]]
            data_rows = rows[1:]

            out_path = self.data_dir / out_name
            with out_path.open("w", encoding="utf-8", newline="") as fh:
                writer = csv.writer(fh)
                writer.writerow(headers)
                for row in data_rows:
                    if not any(cell is not None and str(cell).strip() for cell in row):
                        continue
                    writer.writerow(["" if cell is None else str(cell) for cell in row])

    def _ensure_accounts_json(self) -> None:
        json_path = self.data_dir / "contabilidad_2025_accounts.json"
        if json_path.exists():
            return

        accounts = self._load_accounts()
        with json_path.open("w", encoding="utf-8") as fh:
            json.dump(accounts, fh, ensure_ascii=False, indent=2)

    def _load_accounts(self) -> List[Dict[str, Any]]:
        path = self.data_dir / "contabilidad_2025_accounts.csv"
        rows = []
        with path.open("r", encoding="utf-8-sig", newline="") as fh:
            reader = csv.DictReader(fh)
            for raw in reader:
                code = _clean_str(raw.get("Código") or raw.get("codigo") or raw.get("code"))
                if not code:
                    continue
                rows.append(
                    {
                        "code": code,
                        "description": _clean_str(raw.get("Descripción") or raw.get("descripcion") or raw.get("description")),
                        "group": _clean_str(raw.get("grupo") or raw.get("group")).upper() or "EGRESO",
                        "is_header": _is_truthy(raw.get("es_titular")) or _clean_str(raw.get("tipo")).upper() == "T",
                        "level": int(float(raw.get("nivel") or 0)),
                        "parent_code": _clean_str(raw.get("padre") or raw.get("parent_code")) or None,
                    }
                )
        return rows

    def _load_units(self) -> List[Dict[str, Any]]:
        path = self.data_dir / "contabilidad_2025_unidades_ejecutoras.csv"
        rows = []
        with path.open("r", encoding="utf-8-sig", newline="") as fh:
            reader = csv.DictReader(fh)
            for raw in reader:
                code = _clean_str(raw.get("codigo") or raw.get("code"))
                if not code:
                    continue
                rows.append(
                    {
                        "code": code,
                        "description": _clean_str(raw.get("descripcion") or raw.get("description")),
                        "level": int(float(raw.get("nivel") or 0)),
                        "parent_code": _clean_str(raw.get("padre_codigo") or raw.get("parent_code")) or None,
                    }
                )
        return rows

    def _load_funding_sources(self) -> List[Dict[str, Any]]:
        path = self.data_dir / "contabilidad_2025_fuentes_financiamiento.csv"
        rows = []
        with path.open("r", encoding="utf-8-sig", newline="") as fh:
            reader = csv.DictReader(fh)
            for raw in reader:
                code = _clean_str(raw.get("codigo") or raw.get("code"))
                if not code:
                    continue
                rows.append(
                    {
                        "code": code,
                        "description": _clean_str(raw.get("descripcion") or raw.get("description")),
                    }
                )
        return rows

    def _load_budget_categories(self) -> List[Dict[str, Any]]:
        path = self.data_dir / "contabilidad_2025_categoria_presupuestaria.csv"
        rows = []
        with path.open("r", encoding="utf-8-sig", newline="") as fh:
            reader = csv.DictReader(fh)
            for raw in reader:
                code = _clean_str(raw.get("codigo") or raw.get("code"))
                if not code:
                    continue
                rows.append(
                    {
                        "code": code,
                        "description": _clean_str(raw.get("descripcion") or raw.get("description")),
                    }
                )
        return rows


def _build_tree(items: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    by_code: Dict[str, Dict[str, Any]] = {}
    roots: List[Dict[str, Any]] = []

    for item in items:
        copy = {**item, "children": []}
        by_code[item["code"]] = copy

    for item in by_code.values():
        parent_code = item.get("parent_code")
        if parent_code and parent_code in by_code:
            by_code[parent_code]["children"].append(item)
        else:
            roots.append(item)

    return sorted(roots, key=lambda x: x.get("code", ""))
